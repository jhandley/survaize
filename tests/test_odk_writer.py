from pathlib import Path

import openpyxl

from survaize.model.questionnaire import (
    Option,
    Question,
    Questionnaire,
    QuestionType,
    Section,
    SingleChoiceQuestion,
    TextQuestion,
)
from survaize.writer.odk_writer import ODKWriter


def sample_questionnaire() -> Questionnaire:
    questions: list[Question] = [
        TextQuestion(
            number="1",
            id="name",
            text="What is your name?",
            type=QuestionType.TEXT,
            instructions=None,
            universe=None,
            max_length=None,
        ),
        SingleChoiceQuestion(
            number="2",
            id="gender",
            text="Gender?",
            type=QuestionType.SINGLE_SELECT,
            options=[Option(code="1", label="Male"), Option(code="2", label="Female")],
            instructions=None,
            universe=None,
        ),
    ]

    section = Section(
        id="demo",
        number="A",
        title="Demographics",
        description=None,
        universe=None,
        questions=questions,
        occurrences=1,
    )

    return Questionnaire(title="Test", description=None, id_fields=["name"], sections=[section])


def test_odk_writer_creates_xlsform(tmp_path: Path) -> None:
    questionnaire = sample_questionnaire()
    output_file = tmp_path / "survey.xlsx"

    writer = ODKWriter()
    writer.write(questionnaire, output_file)

    assert output_file.exists()

    workbook = openpyxl.load_workbook(output_file)
    assert "survey" in workbook.sheetnames
    assert "choices" in workbook.sheetnames
    survey_ws = workbook["survey"]
    headers = [cell.value for cell in survey_ws[1]]
    assert headers[:3] == ["type", "name", "label"]
    assert survey_ws.max_row > 2


def test_odk_writer_handles_repeat_sections(tmp_path: Path) -> None:
    """Sections with occurrences > 1 should become repeat groups."""
    questionnaire = sample_questionnaire()
    questionnaire.sections[0].occurrences = 2
    output_file = tmp_path / "repeat.xlsx"

    writer = ODKWriter()
    writer.write(questionnaire, output_file)

    workbook = openpyxl.load_workbook(output_file)
    survey_ws = workbook["survey"]

    first_row = [cell.value for cell in survey_ws[2]]
    assert first_row[0] == "begin repeat"
    assert first_row[1] == "demo"

    last_row = [cell.value for cell in survey_ws[survey_ws.max_row]]
    assert last_row[0] == "end repeat"
    assert last_row[1] == "demo"
